from __future__ import annotations

import os
import secrets
import sqlite3
from contextlib import closing
import psycopg2
from psycopg2.extras import RealDictCursor
from datetime import date, datetime, timedelta
from functools import wraps
from pathlib import Path
from typing import Any

from flask import Flask, flash, g, redirect, render_template, request, send_file, send_from_directory, session, url_for
import cloudinary
import cloudinary.uploader
from openpyxl import Workbook, load_workbook
import io
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

BASE_DIR = Path(__file__).resolve().parent
DATABASE_URL = os.environ.get("DATABASE_URL")
UPLOAD_FOLDER = BASE_DIR / "uploads"
ALLOWED_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "doc", "docx", "xlsx"}
ALLOWED_IMAGE_EXTENSIONS = {"png", "jpg", "jpeg"}

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "change-this-secret-key")
app.config["UPLOAD_FOLDER"] = str(UPLOAD_FOLDER)
UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)

CLOUDINARY_CLOUD_NAME = os.environ.get("CLOUDINARY_CLOUD_NAME", "").strip()
CLOUDINARY_API_KEY = os.environ.get("CLOUDINARY_API_KEY", "").strip()
CLOUDINARY_API_SECRET = os.environ.get("CLOUDINARY_API_SECRET", "").strip()
CLOUDINARY_ENABLED = all([CLOUDINARY_CLOUD_NAME, CLOUDINARY_API_KEY, CLOUDINARY_API_SECRET])

if CLOUDINARY_ENABLED:
    cloudinary.config(
        cloud_name=CLOUDINARY_CLOUD_NAME,
        api_key=CLOUDINARY_API_KEY,
        api_secret=CLOUDINARY_API_SECRET,
        secure=True,
    )

if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL is not set")


class _LastInsertRow:
    def __init__(self, value: int | None):
        self._value = value

    def fetchone(self):
        return {"id": self._value}

    def fetchall(self):
        return [{"id": self._value}] if self._value is not None else []


class PGCursorWrapper:
    def __init__(self, cursor):
        self.cursor = cursor

    def fetchone(self):
        try:
            return self.cursor.fetchone()
        except Exception:
            return None

    def fetchall(self):
        try:
            return self.cursor.fetchall()
        except Exception:
            return []

    def __iter__(self):
        try:
            return iter(self.cursor)
        except Exception:
            return iter(())


class PGConnectionWrapper:
    def __init__(self, conn):
        self.conn = conn
        self.last_insert_id: int | None = None

    def _translate_query(self, query: str) -> str:
        q = query.replace("%", "%%")
        q = q.replace("?", "%s")
        q = q.replace("ifnull(", "COALESCE(")
        return q

    def execute(self, query, params=()):
        normalized = query.strip()
        if normalized.lower().startswith("select last_insert_rowid()"):
            return _LastInsertRow(self.last_insert_id)

        q = self._translate_query(query)
        cur = self.conn.cursor(cursor_factory=RealDictCursor)

        if normalized.lower().startswith("insert into ") and " returning " not in normalized.lower():
            q = q.rstrip().rstrip(";") + " RETURNING id"
            cur.execute(q, params)
            inserted = cur.fetchone()
            self.last_insert_id = inserted["id"] if inserted else None
            return PGCursorWrapper(cur)

        cur.execute(q, params)
        return PGCursorWrapper(cur)

    def executemany(self, query, seq_of_params):
        cur = self.conn.cursor()
        cur.executemany(self._translate_query(query), seq_of_params)
        return cur

    def executescript(self, script):
        cur = self.conn.cursor()
        cur.execute(script)
        return cur

    def commit(self):
        self.conn.commit()

    def rollback(self):
        self.conn.rollback()

    def close(self):
        self.conn.close()


def get_raw_connection():
    return psycopg2.connect(DATABASE_URL)


def is_external_file(value: str | None) -> bool:
    return bool(value) and (str(value).startswith("http://") or str(value).startswith("https://"))


def file_url(value: str | None) -> str | None:
    if not value:
        return None
    return value if is_external_file(value) else url_for("uploaded_file", filename=value)


def upload_file_storage(file, folder: str, allowed_extensions: set[str] | None = None) -> str:
    filename = secure_filename(file.filename or "")
    if not filename:
        raise ValueError("Selected file is invalid.")
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    if allowed_extensions is not None and ext not in allowed_extensions:
        raise ValueError("Unsupported file type.")

    if CLOUDINARY_ENABLED:
        resource_type = "image" if ext in ALLOWED_IMAGE_EXTENSIONS else "raw"
        public_id = f"{folder}/{datetime.now().strftime('%Y%m%d%H%M%S')}_{secrets.token_hex(6)}_{Path(filename).stem}"
        result = cloudinary.uploader.upload(
            file,
            folder=None,
            public_id=public_id,
            resource_type=resource_type,
            overwrite=False,
            use_filename=False,
            unique_filename=False,
        )
        return result.get("secure_url") or result.get("url")

    stored_name = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{secrets.token_hex(4)}_{filename}"
    file_path = UPLOAD_FOLDER / stored_name
    file.save(file_path)
    return stored_name


def delete_stored_file(value: str | None) -> None:
    if not value:
        return
    if is_external_file(value):
        return
    file_path = UPLOAD_FOLDER / value
    if file_path.exists():
        try:
            file_path.unlink()
        except OSError:
            pass


@app.context_processor
def inject_template_helpers():
    return {
        "file_url": file_url,
        "is_external_file": is_external_file,
        "cloudinary_enabled": CLOUDINARY_ENABLED,
    }


# Role list for dropdowns (employee create/edit)
def get_role_options() -> list[str]:
    return [
        "employee",
        "engineer",
        "site_engineer",
        "safety_officer",
        "civil_engineer",
        "electrical_engineer",
        "mechanical_engineer",
        "department_engineer",
        "site_manager",
        "project_engineer",
        "project_manager",
        "manager",
        "hr",
        "admin",
        "super_admin"
    ]


def get_db():
    if "db" not in g:
        raw_conn = get_raw_connection()
        g.db = PGConnectionWrapper(raw_conn)
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


ADMIN_ROLES = {"admin", "super_admin"}
HR_ROLES = {"hr"}
PROJECT_LEAD_ROLES = {"manager", "project_manager", "site_manager"}
TEAM_ACCESS_ROLES = ADMIN_ROLES | HR_ROLES | PROJECT_LEAD_ROLES


def is_admin_role(role: str | None) -> bool:
    return role in ADMIN_ROLES


def is_hr_role(role: str | None) -> bool:
    return role in HR_ROLES


def is_project_scoped_role(role: str | None) -> bool:
    return role in PROJECT_LEAD_ROLES


def can_manage_people(role: str | None) -> bool:
    return role in TEAM_ACCESS_ROLES


def role_label(role: str | None) -> str:
    labels = {
        "admin": "Super Admin",
        "super_admin": "Super Admin",
        "manager": "Manager",
        "project_manager": "Project Manager",
        "site_manager": "Site Manager",
        "engineer": "Engineer",
        "employee": "Employee",
        "hr": "HR",
    }
    return labels.get(role or "", (role or "").replace("_", " ").title())


def visible_user_filter(user: sqlite3.Row, alias: str = "u") -> tuple[str, list[Any]]:
    role = user["role"]
    if is_admin_role(role) or is_hr_role(role):
        return "1=1", []
    if is_project_scoped_role(role):
        project_id = user["project_id"]
        if project_id:
            return f"{alias}.project_id = ?", [project_id]
        return "0=1", []
    return f"{alias}.id = ?", [user["id"]]


def visible_project_filter(user: sqlite3.Row, alias: str = "p") -> tuple[str, list[Any]]:
    role = user["role"]
    if is_admin_role(role) or is_hr_role(role):
        return "1=1", []
    if user["project_id"]:
        return f"{alias}.id = ?", [user["project_id"]]
    return "0=1", []


def user_can_view_employee(viewer: sqlite3.Row, employee: sqlite3.Row) -> bool:
    if viewer["id"] == employee["id"]:
        return True
    if is_admin_role(viewer["role"]) or is_hr_role(viewer["role"]):
        return True
    if is_project_scoped_role(viewer["role"]) and viewer["project_id"] and viewer["project_id"] == employee["project_id"]:
        return True
    return False


def project_choice_rows(db: sqlite3.Connection, user: sqlite3.Row | None = None):
    query = "SELECT * FROM projects"
    params: tuple[Any, ...] = ()
    if user is not None:
        cond, cond_params = visible_project_filter(user, "projects")
        if cond != "1=1":
            query += f" WHERE {cond}"
            params = tuple(cond_params)
    query += " ORDER BY project_name"
    return db.execute(query, params).fetchall()


def team_user_rows(db: sqlite3.Connection, viewer: sqlite3.Row):
    cond, params = visible_user_filter(viewer, "u")
    query = """
        SELECT u.*, d.name AS department_name, ds.name AS designation_name, p.project_name, p.project_code
        FROM users u
        LEFT JOIN departments d ON u.department_id=d.id
        LEFT JOIN designations ds ON u.designation_id=ds.id
        LEFT JOIN projects p ON u.project_id=p.id
        WHERE u.is_active=1 AND """ + cond + " ORDER BY u.full_name"
    return db.execute(query, tuple(params)).fetchall()


def employee_directory_rows(db: sqlite3.Connection, viewer: sqlite3.Row, search: str = "", project_filter: int | None = None):
    query = """
        SELECT u.*, d.name AS department_name, ds.name AS designation_name, p.project_name, p.project_code
        FROM users u
        LEFT JOIN departments d ON u.department_id=d.id
        LEFT JOIN designations ds ON u.designation_id=ds.id
        LEFT JOIN projects p ON u.project_id=p.id
    """
    conditions: list[str] = []
    params: list[Any] = []

    if is_project_scoped_role(viewer["role"]):
        if viewer["project_id"]:
            conditions.append("(u.project_id = ? OR u.project_id IS NULL)")
            params.append(viewer["project_id"])
        else:
            conditions.append("u.project_id IS NULL")
    else:
        visible_condition, visible_params = visible_user_filter(viewer, "u")
        conditions.append(visible_condition)
        params.extend(visible_params)

    if project_filter is not None:
        if project_filter:
            conditions.append("u.project_id = ?")
            params.append(project_filter)
        else:
            conditions.append("u.project_id IS NULL")
    if search:
        conditions.append("(u.full_name LIKE ? OR u.employee_code LIKE ? OR u.email LIKE ? OR ifnull(p.project_name,'') LIKE ?)")
        like = f"%{search}%"
        params.extend([like, like, like, like])
    query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY COALESCE(p.project_name, 'Unassigned'), u.full_name"
    return db.execute(query, tuple(params)).fetchall()


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def calculate_hours(check_in: str | None, check_out: str | None) -> float:
    if not check_in or not check_out:
        return 0.0
    try:
        start = datetime.strptime(check_in, "%H:%M")
        end = datetime.strptime(check_out, "%H:%M")
        diff = (end - start).total_seconds() / 3600
        return round(max(diff, 0), 2)
    except Exception:
        return 0.0



def initialize_postgres() -> None:
    schema = BASE_DIR / "schema.sql"
    if not schema.exists():
        raise RuntimeError(f"schema.sql not found at {schema}")

    with closing(get_raw_connection()) as raw_conn:
        db = PGConnectionWrapper(raw_conn)
        with open(schema, "r", encoding="utf-8") as f:
            db.executescript(f.read())
        db.commit()

        existing = db.execute("SELECT COUNT(*) AS c FROM users").fetchone()
        if not existing or existing["c"] == 0:
            seed_data(db)
            db.commit()


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def allowed_image_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS


def current_user() -> sqlite3.Row | None:
    user_id = session.get("user_id")
    if not user_id:
        return None
    return get_db().execute(
        """
        SELECT u.*, d.name AS department_name, ds.name AS designation_name, m.full_name AS manager_name,
               p.project_name, p.project_code
        FROM users u
        LEFT JOIN departments d ON u.department_id = d.id
        LEFT JOIN designations ds ON u.designation_id = ds.id
        LEFT JOIN users m ON u.manager_id = m.id
        LEFT JOIN projects p ON u.project_id = p.id
        WHERE u.id = ?
        """,
        (user_id,),
    ).fetchone()


def login_required(view):
    @wraps(view)
    def wrapped_view(**kwargs):
        if current_user() is None:
            return redirect(url_for("login"))
        return view(**kwargs)

    return wrapped_view


def role_required(*roles: str):
    def decorator(view):
        @wraps(view)
        def wrapped_view(**kwargs):
            user = current_user()
            if user is None:
                return redirect(url_for("login"))
            if user["role"] not in roles:
                flash("You do not have access to this page.", "danger")
                return redirect(url_for("dashboard"))
            return view(**kwargs)

        return wrapped_view

    return decorator


def log_audit(module_name: str, action_name: str, detail: str, target_user_id: int | None = None) -> None:
    get_db().execute(
        "INSERT INTO audit_logs(actor_user_id, target_user_id, module_name, action_name, detail, created_at) VALUES (?, ?, ?, ?, ?, ?)",
        (session.get("user_id"), target_user_id, module_name, action_name, detail, now_str()),
    )


def create_notification(user_id: int, title: str, message: str, link: str | None = None) -> None:
    get_db().execute(
        "INSERT INTO notifications(user_id, title, message, link, is_read, created_at) VALUES (?, ?, ?, ?, 0, ?)",
        (user_id, title, message, link, now_str()),
    )


def queue_email(to_user_id: int, subject: str, body: str) -> None:
    db = get_db()
    recipient = db.execute("SELECT email FROM users WHERE id = ?", (to_user_id,)).fetchone()
    if recipient:
        db.execute(
            "INSERT INTO email_queue(to_user_id, to_email, subject, body, status, created_at) VALUES (?, ?, ?, ?, 'Queued', ?)",
            (to_user_id, recipient["email"], subject, body, now_str()),
        )


def notify_user(user_id: int, title: str, message: str, link: str | None = None, email_subject: str | None = None, email_body: str | None = None) -> None:
    create_notification(user_id, title, message, link)
    queue_email(user_id, email_subject or title, email_body or message)


def app_counts(user: sqlite3.Row) -> dict[str, Any]:
    db = get_db()
    counts: dict[str, Any] = {}
    counts["my_leave_count"] = db.execute("SELECT COUNT(*) AS c FROM leave_applications WHERE user_id = ?", (user["id"],)).fetchone()["c"]
    counts["my_pending_count"] = db.execute("SELECT COUNT(*) AS c FROM leave_applications WHERE user_id = ? AND status LIKE 'Pending%'", (user["id"],)).fetchone()["c"]
    counts["my_balance"] = db.execute("SELECT COALESCE(SUM(remaining_days),0) AS c FROM leave_balances WHERE user_id = ?", (user["id"],)).fetchone()["c"]
    counts["attendance_days"] = db.execute("SELECT COUNT(*) AS c FROM attendance WHERE user_id = ? AND status='Present'", (user["id"],)).fetchone()["c"]
    counts["payslip_count"] = db.execute("SELECT COUNT(*) AS c FROM payroll_slips WHERE user_id = ?", (user["id"],)).fetchone()["c"]
    counts["unread_notifications"] = db.execute("SELECT COUNT(*) AS c FROM notifications WHERE user_id = ? AND is_read = 0", (user["id"],)).fetchone()["c"]
    counts["project_name"] = user["project_name"] or "Unassigned"
    counts["project_total"] = db.execute("SELECT COUNT(*) AS c FROM projects").fetchone()["c"]
    if is_project_scoped_role(user["role"]):
        counts["team_members"] = db.execute(
            "SELECT COUNT(*) AS c FROM users WHERE is_active = 1 AND project_id = ? AND id <> ?",
            (user["project_id"], user["id"]),
        ).fetchone()["c"] if user["project_id"] else 0
        counts["manager_pending"] = db.execute(
            "SELECT COUNT(*) AS c FROM leave_applications la JOIN users u ON la.user_id = u.id WHERE u.project_id = ? AND la.status LIKE 'Pending%'",
            (user["project_id"],),
        ).fetchone()["c"] if user["project_id"] else 0
    if is_hr_role(user["role"]) or is_admin_role(user["role"]):
        counts["total_employees"] = db.execute("SELECT COUNT(*) AS c FROM users WHERE is_active = 1").fetchone()["c"]
        counts["hr_pending"] = db.execute("SELECT COUNT(*) AS c FROM leave_applications WHERE manager_status='Approved' AND hr_status='Pending'").fetchone()["c"]
        counts["documents_total"] = db.execute("SELECT COUNT(*) AS c FROM employee_documents").fetchone()["c"]
    if is_admin_role(user["role"]):
        counts["queued_emails"] = db.execute("SELECT COUNT(*) AS c FROM email_queue WHERE status = 'Queued'").fetchone()["c"]
    return counts


@app.context_processor
def inject_globals():
    user = current_user()
    settings = None
    unread_count = 0
    if user:
        db = get_db()
        try:
            settings = db.execute("SELECT * FROM company_settings ORDER BY id DESC LIMIT 1").fetchone()
            unread_count = db.execute("SELECT COUNT(*) AS c FROM notifications WHERE user_id=? AND is_read=0", (user["id"],)).fetchone()["c"]
        except sqlite3.OperationalError:
            settings = None
    return {
        "current_user": user,
        "year": date.today().year,
        "settings": settings,
        "unread_notification_count": unread_count,
        "role_label": role_label,
        "is_admin_role": is_admin_role,
        "can_manage_people": can_manage_people,
    }


def seed_data(db=None) -> None:
    own_connection = False
    if db is None:
        db = PGConnectionWrapper(get_raw_connection())
        own_connection = True
    existing_users = db.execute("SELECT COUNT(*) AS c FROM users").fetchone()
    if existing_users and existing_users["c"] > 0:
        if own_connection:
            db.close()
        return

    departments = [
        "Electrical", "Mechanical", "Civil", "Safety", "Operations", "Project Management",
        "HR", "Administration", "Logistics", "Support"
    ]
    designations = [
        "Project Manager", "Operation Manager", "Site Manager", "Site Engineer", "Civil Engineer",
        "Mechanical Engineer", "Safety Engineer", "HR Officer", "System Admin",
        "Civil QC Engineer", "Electrical QC Engineer", "Site Civil Engineer", "Assistant Surveyor",
        "Electrical Technician", "Electrician", "Technician", "Safety Officer", "Driver",
        "Lineman", "Lineman & Store Assistant", "Day Security", "Night Watchman", "Labour", "Tea Boy"
    ]
    for dept in departments:
        db.execute("INSERT INTO departments(name) VALUES (?)", (dept,))
    for desig in designations:
        db.execute("INSERT INTO designations(name) VALUES (?)", (desig,))

    dept_ids = {row["name"]: row["id"] for row in db.execute("SELECT id, name FROM departments").fetchall()}
    desig_ids = {row["name"]: row["id"] for row in db.execute("SELECT id, name FROM designations").fetchall()}

    leave_types = [
        ("Annual Leave", 21, 1),
        ("Sick Leave", 14, 1),
        ("Emergency Leave", 7, 1),
        ("Unpaid Leave", 0, 0),
        ("Casual Leave", 7, 1),
    ]
    db.executemany("INSERT INTO leave_types(name, annual_quota, is_paid) VALUES (?, ?, ?)", leave_types)

    users = [
        {
            "full_name": "Hamid Ali Khan",
            "email": "projectmanager@example.com",
            "employee_code": "PAC-PM-001",
            "password": "Hamid@123",
            "role": "manager",
            "department": "Project Management",
            "designation": "Project Manager",
            "manager_email": None,
            "phone": "+966500000001",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001111",
            "join_date": "2024-01-01",
        },
        {
            "full_name": "Yureed Taseer",
            "email": "opmanager@example.com",
            "employee_code": "PAC-OPS-001",
            "password": "Operation@123",
            "role": "manager",
            "department": "Operations",
            "designation": "Operation Manager",
            "manager_email": "projectmanager@example.com",
            "phone": "+966500000002",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001112",
            "join_date": "2024-02-01",
        },
        {
            "full_name": "Muhammad Waseem",
            "email": "manager@example.com",
            "employee_code": "PAC-249",
            "password": "Muhammad@123",
            "role": "manager",
            "department": "Electrical",
            "designation": "Site Manager",
            "manager_email": "opmanager@example.com",
            "phone": "+966500000003",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001113",
            "join_date": "2024-03-15",
        },
        {
            "full_name": "Faisal Malik",
            "email": "faisal.malik@example.com",
            "employee_code": "PAC-127",
            "password": "Faisal@123",
            "role": "employee",
            "department": "Civil",
            "designation": "Civil QC Engineer",
            "manager_email": "manager@example.com",
            "phone": "+966500000005",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001115",
            "join_date": "2025-02-01",
        },
        {
            "full_name": "Mobeen Ahmad",
            "email": "mobeen.ahmad@example.com",
            "employee_code": "PAC-424",
            "password": "Mobeen@123",
            "role": "employee",
            "department": "Safety",
            "designation": "Safety Officer",
            "manager_email": "manager@example.com",
            "phone": "+966500000006",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001116",
            "join_date": "2025-02-15",
        },
        {
            "full_name": "Qazi Ehsan ul Haq Budder",
            "email": "ehsan.budder@example.com",
            "employee_code": "PAC-450",
            "password": "Qazi@123",
            "role": "employee",
            "department": "Civil",
            "designation": "Site Civil Engineer",
            "manager_email": "manager@example.com",
            "phone": "+966500000010",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001120",
            "join_date": "2025-03-01",
        },
        {
            "full_name": "Usama Jaleel",
            "email": "usama.jaleel@example.com",
            "employee_code": "PAC-452",
            "password": "Usama@123",
            "role": "employee",
            "department": "Civil",
            "designation": "Assistant Surveyor",
            "manager_email": "manager@example.com",
            "phone": "+966500000011",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001121",
            "join_date": "2025-03-05",
        },
        {
            "full_name": "Abid Ali",
            "email": "abid.ali@example.com",
            "employee_code": "PAC-455",
            "password": "Abid@123",
            "role": "employee",
            "department": "Electrical",
            "designation": "Electrical Technician",
            "manager_email": "manager@example.com",
            "phone": "+966500000012",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001122",
            "join_date": "2025-03-10",
        },
        {
            "full_name": "Shabir Khan",
            "email": "shabir.khan@example.com",
            "employee_code": "PAC-421",
            "password": "Shabir@123",
            "role": "employee",
            "department": "Administration",
            "designation": "Night Watchman",
            "manager_email": "manager@example.com",
            "phone": "+966500000013",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001123",
            "join_date": "2025-03-10",
        },
        {
            "full_name": "Sagir Mridha",
            "email": "sagir.mridha@example.com",
            "employee_code": "PAC-457",
            "password": "Sagir@123",
            "role": "employee",
            "department": "Logistics",
            "designation": "Driver",
            "manager_email": "manager@example.com",
            "phone": "+966500000014",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001124",
            "join_date": "2025-03-12",
        },
        {
            "full_name": "Ashfaq Ahmad",
            "email": "ashfaq.ahmad@example.com",
            "employee_code": "PAC-295",
            "password": "Ashfaq@123",
            "role": "employee",
            "department": "Electrical",
            "designation": "Electrical QC Engineer",
            "manager_email": "manager@example.com",
            "phone": "+966500000015",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001125",
            "join_date": "2025-03-15",
        },
        {
            "full_name": "Vijendra Singh",
            "email": "vijendra.singh@example.com",
            "employee_code": "PAC-462",
            "password": "Vijendra@123",
            "role": "employee",
            "department": "Electrical",
            "designation": "Electrician",
            "manager_email": "manager@example.com",
            "phone": "+966500000016",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001126",
            "join_date": "2025-03-18",
        },
        {
            "full_name": "Salahudddin",
            "email": "salahudddin@example.com",
            "employee_code": "PAC-313",
            "password": "Salahuddin@123",
            "role": "employee",
            "department": "Safety",
            "designation": "Safety Engineer",
            "manager_email": "manager@example.com",
            "phone": "+966500000017",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001127",
            "join_date": "2025-03-18",
        },
        {
            "full_name": "Mirza Samiulla Baig",
            "email": "mirza.samiulla@example.com",
            "employee_code": "PAC-224",
            "password": "Mirza@123",
            "role": "employee",
            "department": "Mechanical",
            "designation": "Mechanical Engineer",
            "manager_email": "manager@example.com",
            "phone": "+966500000018",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001128",
            "join_date": "2025-03-20",
        },
        {
            "full_name": "Muhammad Ashfaq",
            "email": "muhammad.ashfaq@example.com",
            "employee_code": "PAC-324",
            "password": "MuhammadAshfaq@123",
            "role": "employee",
            "department": "Mechanical",
            "designation": "Technician",
            "manager_email": "manager@example.com",
            "phone": "+966500000019",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001129",
            "join_date": "2025-03-22",
        },
        {
            "full_name": "Sayem",
            "email": "sayem@example.com",
            "employee_code": "PAC-476",
            "password": "Sayem@123",
            "role": "employee",
            "department": "Support",
            "designation": "Tea Boy",
            "manager_email": "manager@example.com",
            "phone": "+966500000020",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001130",
            "join_date": "2025-03-22",
        },
        {
            "full_name": "Suhel Prdhan",
            "email": "suhel.prdhan@example.com",
            "employee_code": "PAC-190",
            "password": "Suhel@123",
            "role": "employee",
            "department": "Support",
            "designation": "Labour",
            "manager_email": "manager@example.com",
            "phone": "+966500000021",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001131",
            "join_date": "2025-03-22",
        },
        {
            "full_name": "Adnan Yaqob",
            "email": "adnan.yaqob@example.com",
            "employee_code": "PAC-376",
            "password": "Adnan@123",
            "role": "employee",
            "department": "Electrical",
            "designation": "Lineman",
            "manager_email": "manager@example.com",
            "phone": "+966500000022",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001132",
            "join_date": "2025-03-22",
        },
        {
            "full_name": "Umair Ali",
            "email": "umair.ali@example.com",
            "employee_code": "PAC-378",
            "password": "Umair@123",
            "role": "employee",
            "department": "Electrical",
            "designation": "Lineman & Store Assistant",
            "manager_email": "manager@example.com",
            "phone": "+966500000023",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001133",
            "join_date": "2025-03-22",
        },
        {
            "full_name": "Ghanem Alsqour",
            "email": "ghanem.alsqour@example.com",
            "employee_code": "PAC-464",
            "password": "Ghanem@123",
            "role": "employee",
            "department": "Administration",
            "designation": "Day Security",
            "manager_email": "manager@example.com",
            "phone": "+966500000024",
            "address": "Dammam, Saudi Arabia",
            "emergency_contact": "+966500001134",
            "join_date": "2025-03-22",
        },
        {
            "full_name": "Asif Chandal",
            "email": "hr@example.com",
            "employee_code": "EMP1008",
            "password": "HR@12345",
            "role": "hr",
            "department": "HR",
            "designation": "HR Officer",
            "manager_email": None,
            "phone": "+966500000008",
            "address": "Khobar, Saudi Arabia",
            "emergency_contact": "+966500001118",
            "join_date": "2023-08-10",
        },
        {
            "full_name": "Super Admin",
            "email": "admin@example.com",
            "employee_code": "EMP1009",
            "password": "Admin@123",
            "role": "admin",
            "department": "Administration",
            "designation": "System Admin",
            "manager_email": None,
            "phone": "+966500000009",
            "address": "Riyadh, Saudi Arabia",
            "emergency_contact": "+966500001119",
            "join_date": "2023-01-01",
        },
    ]

    user_id_by_email = {}
    for row in users:
        manager_id = user_id_by_email.get(row["manager_email"])
        db.execute(
            """
            INSERT INTO users(full_name, email, employee_code, password_hash, role, department_id, designation_id, manager_id, phone, address, emergency_contact, join_date, is_active)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
            """,
            (
                row["full_name"],
                row["email"],
                row["employee_code"],
                generate_password_hash(row["password"]),
                row["role"],
                dept_ids[row["department"]],
                desig_ids[row["designation"]],
                manager_id,
                row["phone"],
                row["address"],
                row["emergency_contact"],
                row["join_date"],
            ),
        )
        user_id_by_email[row["email"]] = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]

    for user_id in user_id_by_email.values():
        for leave_type_id, leave_def in enumerate(leave_types, start=1):
            default_total = leave_def[1]
            db.execute(
                "INSERT INTO leave_balances(user_id, leave_type_id, total_days, used_days, remaining_days) VALUES (?, ?, ?, 0, ?)",
                (user_id, leave_type_id, default_total, default_total),
            )

    db.execute(
        "INSERT INTO company_settings(company_name, leave_workflow, default_working_hours, allow_document_upload) VALUES (?, ?, ?, ?)",
        ("Pacost International", "Site Engineer / Site Staff → Site Manager → HR Final Review", 8.0, 1),
    )

    created_at = now_str()
    site_engineer_id = user_id_by_email["faisal.malik@example.com"]
    site_manager_id = user_id_by_email["manager@example.com"]
    hr_id = user_id_by_email["hr@example.com"]
    admin_id = user_id_by_email["admin@example.com"]

    db.execute(
        """
        INSERT INTO leave_applications(application_no, user_id, leave_type_id, from_date, to_date, total_days, reason, attachment, status, manager_status, hr_status, current_stage, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        ("LV-2026-0001", site_engineer_id, 1, "2026-03-25", "2026-03-27", 3, "Family commitment", None, "Pending HR Approval", "Approved", "Pending", "hr_review", created_at),
    )
    leave_id = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
    db.executemany(
        "INSERT INTO leave_history(leave_application_id, action, remarks, action_by, action_at) VALUES (?, ?, ?, ?, ?)",
        [
            (leave_id, "Submitted", "Employee submitted application", site_engineer_id, created_at),
            (leave_id, "Manager Approved", "Approved by Muhammad Waseem", site_manager_id, created_at),
        ],
    )

    attendance_seed_users = list(user_id_by_email.values())
    for uid in attendance_seed_users:
        for day_offset in range(0, 10):
            dt = date.today() - timedelta(days=day_offset)
            if dt.weekday() >= 5:
                status = "Weekend"
                cin = None
                cout = None
                hours = 0
            else:
                status = "Present"
                cin = "08:00"
                cout = "17:00"
                hours = 8
            db.execute(
                "INSERT INTO attendance(user_id, attendance_date, check_in, check_out, status, hours_worked, remarks) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (uid, dt.isoformat(), cin, cout, status, hours, None),
            )

    payslips = [
        (site_manager_id, "Feb 2026", 9500, 1500, 300, 10700),
        (site_engineer_id, "Feb 2026", 4500, 500, 100, 4900),
        (user_id_by_email["faisal.malik@example.com"], "Feb 2026", 6000, 800, 150, 6650),
        (user_id_by_email["mirza.samiulla@example.com"], "Feb 2026", 6200, 900, 200, 6900),
        (user_id_by_email["salahudddin@example.com"], "Feb 2026", 5800, 700, 120, 6380),
    ]
    for row in payslips:
        db.execute(
            "INSERT INTO payroll_slips(user_id, month_label, basic_salary, allowances, deductions, net_salary, generated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (*row, created_at),
        )

    db.executemany(
        "INSERT INTO audit_logs(actor_user_id, target_user_id, module_name, action_name, detail, created_at) VALUES (?, ?, ?, ?, ?, ?)",
        [
            (admin_id, site_engineer_id, "Payroll", "Seeded", "Created demo payroll slips", created_at),
            (hr_id, site_engineer_id, "Leave", "Reviewed", "HR review pending for demo request", created_at),
        ],
    )

    notifications = [
        (site_engineer_id, "Leave request update", "Your annual leave request is pending HR review.", "/leaves", 0, created_at),
        (site_manager_id, "Team leave awaiting attention", "A leave request has already moved to HR stage.", "/leaves", 0, created_at),
        (admin_id, "System ready", "Demo data was created successfully.", "/settings", 0, created_at),
    ]
    db.executemany("INSERT INTO notifications(user_id, title, message, link, is_read, created_at) VALUES (?, ?, ?, ?, ?, ?)", notifications)
    email_seed = [
        (site_engineer_id, "faisal.malik@example.com", "Leave request received", "Your leave request LV-2026-0001 is now pending HR review.", "Queued", created_at),
        (admin_id, "admin@example.com", "Demo portal initialized", "The portal database and demo users are ready.", "Queued", created_at),
    ]
    db.executemany("INSERT INTO email_queue(to_user_id, to_email, subject, body, status, created_at) VALUES (?, ?, ?, ?, ?, ?)", email_seed)

    db.execute("UPDATE users SET monthly_basic=?, default_allowances=?, deduction_per_absent=?, deduction_per_late=? WHERE email=?", (9500, 1500, 316.67, 50, "manager@example.com"))
    db.execute("UPDATE users SET monthly_basic=?, default_allowances=?, deduction_per_absent=?, deduction_per_late=? WHERE employee_code=?", (9500, 1500, 316.67, 50, "PAC-249"))
    db.commit()
    if own_connection:
        db.close()


@app.route("/")
def index():
    return redirect(url_for("dashboard" if current_user() else "login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        identifier = (request.form.get("identifier") or "").strip()
        password = request.form.get("password") or ""
        if not identifier or not password:
            flash("Please enter both your email or employee code and password.", "danger")
            return render_template("login.html")
        db = get_db()
        user = db.execute(
            "SELECT * FROM users WHERE lower(email) = lower(?) OR upper(employee_code) = upper(?)",
            (identifier, identifier),
        ).fetchone()
        if user and check_password_hash(user["password_hash"], password) and user["is_active"]:
            session.clear()
            session["user_id"] = user["id"]
            flash(f"Welcome back, {user['full_name']}.", "success")
            return redirect(url_for("dashboard"))
        flash("Invalid login credentials.", "danger")
    return render_template("login.html")


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE lower(email) = ?", (email,)).fetchone()
        if user and user["is_active"]:
            temp_password = f"Temp@{secrets.token_hex(4)}A1"
            db.execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(temp_password), user["id"]))
            db.execute(
                "INSERT INTO email_queue(to_user_id, to_email, subject, body, status, created_at) VALUES (?, ?, ?, ?, 'Queued', ?)",
                (user["id"], user["email"], "Password reset", f"Your temporary password is: {temp_password}\nPlease log in and change it immediately.", now_str()),
            )
            db.execute(
                "INSERT INTO notifications(user_id, title, message, link, is_read, created_at) VALUES (?, ?, ?, ?, 0, ?)",
                (user["id"], "Password reset requested", "A temporary password has been generated and queued in the email center.", url_for("change_password"), now_str()),
            )
            db.commit()
        flash("If that email exists, a temporary password has been generated and queued.", "info")
        return redirect(url_for("login"))
    return render_template("forgot_password.html")


@app.route("/logout")
@login_required
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    user = current_user()
    db = get_db()
    counts = app_counts(user)
    if user["role"] == "employee":
        recent_leaves = db.execute("SELECT la.*, lt.name AS leave_type_name FROM leave_applications la JOIN leave_types lt ON la.leave_type_id=lt.id WHERE la.user_id=? ORDER BY la.id DESC LIMIT 5", (user["id"],)).fetchall()
    elif is_project_scoped_role(user["role"]):
        recent_leaves = db.execute(
            "SELECT la.*, lt.name AS leave_type_name, u.full_name FROM leave_applications la JOIN leave_types lt ON la.leave_type_id=lt.id JOIN users u ON la.user_id=u.id WHERE u.project_id=? ORDER BY la.id DESC LIMIT 8",
            (user["project_id"],),
        ).fetchall() if user["project_id"] else []
    else:
        recent_leaves = db.execute("SELECT la.*, lt.name AS leave_type_name, u.full_name FROM leave_applications la JOIN leave_types lt ON la.leave_type_id=lt.id JOIN users u ON la.user_id=u.id ORDER BY la.id DESC LIMIT 8").fetchall()
    recent_attendance = db.execute("SELECT * FROM attendance WHERE user_id=? ORDER BY attendance_date DESC LIMIT 5", (user["id"],)).fetchall()
    recent_notifications = db.execute("SELECT * FROM notifications WHERE user_id=? ORDER BY id DESC LIMIT 5", (user["id"],)).fetchall()
    project_rows = project_choice_rows(db, user)
    return render_template("dashboard.html", counts=counts, recent_leaves=recent_leaves, recent_attendance=recent_attendance, recent_notifications=recent_notifications, project_rows=project_rows)


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    user = current_user()
    db = get_db()
    if request.method == "POST":
        phone = (request.form.get("phone") or "").strip()
        address = (request.form.get("address") or "").strip()
        emergency_contact = (request.form.get("emergency_contact") or "").strip()
        avatar = request.files.get("profile_picture")
        avatar_filename = user["avatar_filename"]
        if avatar and avatar.filename:
            if allowed_image_file(avatar.filename):
                ext = avatar.filename.rsplit('.', 1)[1].lower()
                old_avatar = user["avatar_filename"]
                avatar_filename = upload_file_storage(avatar, "employee_portal/profile_pictures", ALLOWED_IMAGE_EXTENSIONS)
                if old_avatar and old_avatar != avatar_filename:
                    delete_stored_file(old_avatar)
            else:
                flash("Profile picture must be a PNG, JPG, or JPEG image.", "warning")
                return redirect(url_for("profile"))
        db.execute(
            "UPDATE users SET phone=?, address=?, emergency_contact=?, avatar_filename=? WHERE id=?",
            (phone, address, emergency_contact, avatar_filename, user["id"]),
        )
        log_audit("Profile", "Updated", "Updated own contact details", user["id"])
        db.commit()
        flash("Profile updated successfully.", "success")
        return redirect(url_for("profile"))
    balances = db.execute("SELECT lb.*, lt.name AS leave_type_name FROM leave_balances lb JOIN leave_types lt ON lb.leave_type_id=lt.id WHERE lb.user_id=? ORDER BY lt.name", (user["id"],)).fetchall()
    return render_template("profile.html", balances=balances)


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    user = current_user()
    db = get_db()
    if request.method == "POST":
        current_password = request.form["current_password"]
        new_password = request.form["new_password"]
        confirm_password = request.form["confirm_password"]
        if not check_password_hash(user["password_hash"], current_password):
            flash("Current password is not correct.", "danger")
        elif len(new_password) < 8:
            flash("New password must be at least 8 characters.", "danger")
        elif new_password != confirm_password:
            flash("New password and confirmation do not match.", "danger")
        else:
            db.execute("UPDATE users SET password_hash=? WHERE id=?", (generate_password_hash(new_password), user["id"]))
            log_audit("Security", "Password Changed", "Changed own password", user["id"])
            db.commit()
            flash("Password changed successfully.", "success")
            return redirect(url_for("dashboard"))
    return render_template("change_password.html")


@app.route("/leave/apply", methods=["GET", "POST"])
@login_required
def apply_leave():
    db = get_db()
    user = current_user()
    leave_types = db.execute("SELECT * FROM leave_types ORDER BY name").fetchall()
    if request.method == "POST":
        leave_type_id = int(request.form["leave_type_id"])
        from_date = request.form["from_date"]
        to_date = request.form["to_date"]
        reason = request.form["reason"].strip()
        start = datetime.strptime(from_date, "%Y-%m-%d").date()
        end = datetime.strptime(to_date, "%Y-%m-%d").date()
        total_days = (end - start).days + 1
        if total_days <= 0:
            flash("To date must be on or after from date.", "danger")
            return render_template("apply_leave.html", leave_types=leave_types)
        attachment_name = None
        file = request.files.get("attachment")
        if file and file.filename:
            if not allowed_file(file.filename):
                flash("Unsupported file type.", "danger")
                return render_template("apply_leave.html", leave_types=leave_types)
            attachment_name = upload_file_storage(file, "leave_attachments", ALLOWED_EXTENSIONS)
        next_num = db.execute("SELECT COUNT(*) AS c FROM leave_applications").fetchone()["c"] + 1
        app_no = f"LV-2026-{next_num:04d}"
        db.execute(
            "INSERT INTO leave_applications(application_no, user_id, leave_type_id, from_date, to_date, total_days, reason, attachment, status, manager_status, hr_status, current_stage, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (app_no, user["id"], leave_type_id, from_date, to_date, total_days, reason, attachment_name, "Pending Manager Approval", "Pending", "Pending", "manager_review", now_str()),
        )
        leave_id = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        db.execute("INSERT INTO leave_history(leave_application_id, action, remarks, action_by, action_at) VALUES (?, ?, ?, ?, ?)", (leave_id, "Submitted", "Employee submitted leave request", user["id"], now_str()))
        if user["manager_id"]:
            notify_user(
                user["manager_id"],
                "New leave request",
                f"{user['full_name']} submitted leave request {app_no}.",
                link=url_for("leave_detail", leave_id=leave_id),
                email_subject=f"Leave request {app_no}",
                email_body=f"A new leave request from {user['full_name']} requires your review.",
            )
        notify_user(user["id"], "Leave submitted", f"Your leave request {app_no} was submitted successfully.", url_for("my_leaves"))
        log_audit("Leave", "Submitted", f"Leave request {app_no} submitted", user["id"])
        db.commit()
        flash("Leave application submitted successfully.", "success")
        return redirect(url_for("my_leaves"))
    return render_template("apply_leave.html", leave_types=leave_types)


@app.route("/leaves")
@app.route("/leave-tracking")
@login_required
def my_leaves():
    user = current_user()
    db = get_db()
    query = "SELECT la.*, lt.name AS leave_type_name, u.full_name FROM leave_applications la JOIN leave_types lt ON la.leave_type_id=lt.id JOIN users u ON la.user_id=u.id"
    params: tuple[Any, ...] = ()
    if user["role"] == "employee":
        query += " WHERE la.user_id=?"
        params = (user["id"],)
    elif is_project_scoped_role(user["role"]):
        if user["project_id"]:
            query += " WHERE (u.project_id=? OR la.user_id=?)"
            params = (user["project_id"], user["id"])
        else:
            query += " WHERE la.user_id=?"
            params = (user["id"],)
    query += " ORDER BY la.id DESC"
    leaves = db.execute(query, params).fetchall()
    return render_template("leaves.html", leaves=leaves, page_title="Leave Tracking")


@app.route("/leave/<int:leave_id>", methods=["GET", "POST"])
@login_required
def leave_detail(leave_id: int):
    user = current_user()
    db = get_db()
    leave = db.execute(
        "SELECT la.*, lt.name AS leave_type_name, u.full_name, u.manager_id FROM leave_applications la JOIN leave_types lt ON la.leave_type_id=lt.id JOIN users u ON la.user_id=u.id WHERE la.id=?",
        (leave_id,),
    ).fetchone()
    if not leave:
        flash("Leave application not found.", "danger")
        return redirect(url_for("my_leaves"))
    can_view = is_hr_role(user["role"]) or is_admin_role(user["role"]) or leave["user_id"] == user["id"] or (is_project_scoped_role(user["role"]) and user["project_id"] == db.execute("SELECT project_id FROM users WHERE id=?", (leave["user_id"],)).fetchone()["project_id"])
    if not can_view:
        flash("You do not have access to this record.", "danger")
        return redirect(url_for("my_leaves"))
    if request.method == "POST":
        action = request.form["action"]
        remarks = request.form.get("remarks", "").strip() or None
        hist_action = None
        if user["role"] in {"manager", "project_manager", "site_manager"} and leave["manager_id"] == user["id"] and leave["manager_status"] == "Pending":
            if action == "approve":
                db.execute("UPDATE leave_applications SET manager_status='Approved', status='Pending HR Approval', current_stage='hr_review' WHERE id=?", (leave_id,))
                hist_action = "Manager Approved"
                hr_user = db.execute("SELECT id FROM users WHERE role='hr' AND is_active=1 ORDER BY id LIMIT 1").fetchone()
                if hr_user:
                    notify_user(hr_user["id"], "Leave request pending HR review", f"{leave['application_no']} is ready for HR review.", url_for("leave_detail", leave_id=leave_id))
                notify_user(leave["user_id"], "Manager approved leave", f"{leave['application_no']} has moved to HR review.", url_for("leave_detail", leave_id=leave_id))
            elif action == "reject":
                db.execute("UPDATE leave_applications SET manager_status='Rejected', status='Rejected by Manager', current_stage='closed' WHERE id=?", (leave_id,))
                hist_action = "Manager Rejected"
                notify_user(leave["user_id"], "Leave rejected", f"{leave['application_no']} was rejected by your manager.", url_for("leave_detail", leave_id=leave_id))
        elif user["role"] in {"hr", "admin"} and leave["manager_status"] == "Approved" and leave["hr_status"] == "Pending":
            if action == "approve":
                db.execute("UPDATE leave_applications SET hr_status='Approved', status='Final Approved', current_stage='closed' WHERE id=?", (leave_id,))
                db.execute("UPDATE leave_balances SET used_days=used_days+?, remaining_days=remaining_days-? WHERE user_id=? AND leave_type_id=?", (leave["total_days"], leave["total_days"], leave["user_id"], leave["leave_type_id"]))
                hist_action = "HR Approved"
                notify_user(leave["user_id"], "Leave approved", f"{leave['application_no']} was finally approved.", url_for("leave_detail", leave_id=leave_id))
            elif action == "reject":
                db.execute("UPDATE leave_applications SET hr_status='Rejected', status='Rejected by HR', current_stage='closed' WHERE id=?", (leave_id,))
                hist_action = "HR Rejected"
                notify_user(leave["user_id"], "Leave rejected", f"{leave['application_no']} was rejected by HR.", url_for("leave_detail", leave_id=leave_id))
        if hist_action:
            db.execute("INSERT INTO leave_history(leave_application_id, action, remarks, action_by, action_at) VALUES (?, ?, ?, ?, ?)", (leave_id, hist_action, remarks, user["id"], now_str()))
            log_audit("Leave", hist_action, f"Leave request {leave['application_no']} actioned", leave["user_id"])
            db.commit()
            flash("Action saved successfully.", "success")
            return redirect(url_for("leave_detail", leave_id=leave_id))
    history = db.execute("SELECT lh.*, u.full_name FROM leave_history lh LEFT JOIN users u ON lh.action_by=u.id WHERE lh.leave_application_id=? ORDER BY lh.id ASC", (leave_id,)).fetchall()
    return render_template("leave_detail.html", leave=leave, history=history)


@app.route("/team")
@login_required
@role_required("manager", "project_manager", "site_manager", "engineer", "hr", "admin", "super_admin")
def team():
    user = current_user()
    db = get_db()
    search = request.args.get("q", "").strip()
    project_filter = request.args.get("project_id", type=int)
    employees = employee_directory_rows(db, user, search=search, project_filter=project_filter)
    projects = project_choice_rows(db, user)
    return render_template("team.html", employees=employees, search=search, projects=projects, selected_project_id=project_filter)


@app.route("/team/export")
@login_required
@role_required("manager", "project_manager", "site_manager", "engineer", "hr", "admin", "super_admin")
def export_team():
    user = current_user()
    db = get_db()
    search = request.args.get("q", "").strip()
    project_filter = request.args.get("project_id", type=int)
    export_scope = (request.args.get("scope") or "current").strip().lower()

    if export_scope == "full" and not (is_admin_role(user["role"]) or is_hr_role(user["role"])):
        flash("You can export employees from your own project only.", "danger")
        return redirect(url_for("team", q=search, project_id=project_filter))

    effective_project_filter = None if export_scope == "full" else project_filter
    employees = employee_directory_rows(db, user, search=search, project_filter=effective_project_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    headers = [
        "Employee Code",
        "Full Name",
        "Email",
        "Phone",
        "Project Code",
        "Project Name",
        "Department",
        "Designation",
        "Role",
        "Joining Date",
        "Status",
    ]
    ws.append(headers)

    for row in employees:
        ws.append([
            row["employee_code"],
            row["full_name"],
            row["email"],
            row["phone"] or "",
            row["project_code"] or "",
            row["project_name"] or "",
            row["department_name"] or "",
            row["designation_name"] or "",
            role_label(row["role"]),
            row["join_date"] or "",
            "Active" if row["is_active"] else "Inactive",
        ])

    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    widths = {
        "A": 18, "B": 26, "C": 30, "D": 18, "E": 16, "F": 24,
        "G": 18, "H": 20, "I": 18, "J": 16, "K": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"

    meta = wb.create_sheet("Export Info")
    meta.append(["Generated At", now_str()])
    meta.append(["Generated By", user["full_name"]])
    meta.append(["Role", role_label(user["role"])])
    meta.append(["Scope", "Full Employee List" if export_scope == "full" else "Project / Filtered Employee List"])
    meta.append(["Search", search or "All"])
    if effective_project_filter:
        project = db.execute("SELECT project_code, project_name FROM projects WHERE id=?", (effective_project_filter,)).fetchone()
        project_label = f"{project['project_code']} - {project['project_name']}" if project else str(effective_project_filter)
    elif export_scope == "full":
        project_label = "All Projects"
    else:
        project_label = user["project_name"] or "All Visible Projects"
    meta.append(["Project Filter", project_label])
    meta.append(["Total Employees", len(employees)])
    meta.column_dimensions["A"].width = 20
    meta.column_dimensions["B"].width = 40

    filename_scope = "full_employee_list" if export_scope == "full" else "project_employee_list"
    safe_project = project_label.lower().replace(" ", "_").replace("/", "-") if project_label else "all"
    safe_project = "".join(ch for ch in safe_project if ch.isalnum() or ch in {"_", "-"})[:40] or "all"
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"{filename_scope}_{safe_project}_{date.today().isoformat()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/employees/<int:user_id>")
@login_required
def employee_detail(user_id: int):
    viewer = current_user()
    db = get_db()
    employee = db.execute(
        "SELECT u.*, d.name AS department_name, ds.name AS designation_name, m.full_name AS manager_name, p.project_name, p.project_code FROM users u LEFT JOIN departments d ON u.department_id=d.id LEFT JOIN designations ds ON u.designation_id=ds.id LEFT JOIN users m ON u.manager_id=m.id LEFT JOIN projects p ON u.project_id=p.id WHERE u.id=?",
        (user_id,),
    ).fetchone()
    if not employee:
        flash("Employee not found.", "danger")
        return redirect(url_for("team"))
    can_view = user_can_view_employee(viewer, employee)
    if not can_view:
        flash("You do not have access to this employee.", "danger")
        return redirect(url_for("dashboard"))
    balances = db.execute("SELECT lb.*, lt.name AS leave_type_name FROM leave_balances lb JOIN leave_types lt ON lb.leave_type_id=lt.id WHERE lb.user_id=? ORDER BY lt.name", (user_id,)).fetchall()
    attendance_rows = db.execute("SELECT * FROM attendance WHERE user_id=? ORDER BY attendance_date DESC LIMIT 10", (user_id,)).fetchall()
    slips = db.execute("SELECT * FROM payroll_slips WHERE user_id=? ORDER BY id DESC LIMIT 6", (user_id,)).fetchall()
    docs = db.execute("SELECT * FROM employee_documents WHERE user_id=? ORDER BY id DESC LIMIT 6", (user_id,)).fetchall()
    return render_template("employee_detail.html", employee=employee, balances=balances, attendance_rows=attendance_rows, slips=slips, docs=docs)


@app.route("/projects", methods=["GET", "POST"])
@login_required
@role_required("admin", "super_admin")
def projects_view():
    db = get_db()
    edit_id = request.args.get("edit", type=int)
    project = None
    if edit_id:
        project = db.execute("SELECT * FROM projects WHERE id=?", (edit_id,)).fetchone()
    if request.method == "POST":
        project_id = request.form.get("project_id", type=int)
        project_code = (request.form.get("project_code") or "").strip().upper()
        project_name = (request.form.get("project_name") or "").strip()
        location = (request.form.get("location") or "").strip()
        client_name = (request.form.get("client_name") or "").strip()
        status = (request.form.get("status") or "Active").strip()
        start_date = (request.form.get("start_date") or "").strip() or None
        end_date = (request.form.get("end_date") or "").strip() or None
        if not project_code or not project_name:
            flash("Project code and project name are required.", "danger")
        else:
            existing = db.execute("SELECT id FROM projects WHERE project_code=? AND id != ?", (project_code, project_id or 0)).fetchone()
            if existing:
                flash("Project code already exists.", "danger")
            else:
                if project_id:
                    db.execute("UPDATE projects SET project_code=?, project_name=?, location=?, client_name=?, status=?, start_date=?, end_date=? WHERE id=?", (project_code, project_name, location, client_name, status, start_date, end_date, project_id))
                    log_audit("Projects", "Updated", f"Updated project {project_code}")
                    flash("Project updated successfully.", "success")
                else:
                    db.execute("INSERT INTO projects(project_code, project_name, location, client_name, status, start_date, end_date, created_by, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)", (project_code, project_name, location, client_name, status, start_date, end_date, session.get("user_id"), now_str()))
                    log_audit("Projects", "Created", f"Created project {project_code}")
                    flash("Project created successfully.", "success")
                db.commit()
                return redirect(url_for("projects_view"))
    projects = db.execute("SELECT p.*, COUNT(u.id) AS employee_count FROM projects p LEFT JOIN users u ON u.project_id=p.id AND u.is_active=1 GROUP BY p.id ORDER BY p.project_name").fetchall()
    return render_template("projects.html", projects=projects, project=project)


@app.route("/projects/<int:project_id>/delete", methods=["POST"])
@login_required
@role_required("admin", "super_admin")
def delete_project(project_id: int):
    db = get_db()
    project = db.execute("SELECT * FROM projects WHERE id=?", (project_id,)).fetchone()
    if not project:
        flash("Project not found.", "danger")
        return redirect(url_for("projects_view"))

    assigned_count = db.execute(
        "SELECT COUNT(*) AS c FROM users WHERE project_id=?",
        (project_id,),
    ).fetchone()["c"]
    if assigned_count > 0:
        flash("Cannot delete this project because employees are still assigned to it. Reassign them first.", "danger")
        return redirect(url_for("projects_view"))

    total_projects = db.execute("SELECT COUNT(*) AS c FROM projects").fetchone()["c"]
    if total_projects <= 1:
        flash("You cannot delete the last remaining project.", "danger")
        return redirect(url_for("projects_view"))

    db.execute("DELETE FROM projects WHERE id=?", (project_id,))
    log_audit("Projects", "Deleted", f"Deleted project {project['project_code']}")
    db.commit()
    flash("Project deleted successfully.", "success")
    return redirect(url_for("projects_view"))


@app.route("/employees/bulk-upload", methods=["GET", "POST"])
@login_required
@role_required("hr", "admin", "super_admin")
def bulk_employee_upload():
    db = get_db()
    if request.method == "POST":
        upload = request.files.get("file")
        if not upload or not upload.filename:
            flash("Please choose an Excel file first.", "warning")
            return redirect(url_for("bulk_employee_upload"))
        ext = upload.filename.rsplit(".", 1)[-1].lower() if "." in upload.filename else ""
        if ext not in {"xlsx", "xlsm", "xltx", "xltm"}:
            flash("Bulk employee upload supports Excel .xlsx files only.", "warning")
            return redirect(url_for("bulk_employee_upload"))
        try:
            wb = load_workbook(filename=io.BytesIO(upload.read()), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                flash("The uploaded Excel file is empty.", "warning")
                return redirect(url_for("bulk_employee_upload"))
            headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
            required = ["full_name", "email", "employee_code", "password", "role", "department", "designation"]
            missing = [c for c in required if c not in headers]
            if missing:
                flash(f"Missing required columns: {', '.join(missing)}", "danger")
                return redirect(url_for("bulk_employee_upload"))
            idx = {name: headers.index(name) for name in headers if name}

            def cell(row, key, default=""):
                pos = idx.get(key)
                if pos is None or pos >= len(row) or row[pos] is None:
                    return default
                return str(row[pos]).strip()

            created = 0
            skipped = 0
            for row in rows[1:]:
                if not row or not any(v is not None and str(v).strip() for v in row):
                    continue
                full_name = cell(row, "full_name")
                email = cell(row, "email").lower()
                employee_code = cell(row, "employee_code")
                password = cell(row, "password")
                role = cell(row, "role", "employee").lower()
                department_name = cell(row, "department")
                designation_name = cell(row, "designation")
                if not full_name or not email or not employee_code or not password or role not in get_role_options() or not department_name or not designation_name:
                    skipped += 1
                    continue
                existing = db.execute("SELECT id FROM users WHERE lower(email)=? OR employee_code=?", (email, employee_code)).fetchone()
                if existing:
                    skipped += 1
                    continue
                dept = db.execute("SELECT id FROM departments WHERE lower(name)=lower(?)", (department_name,)).fetchone()
                if not dept:
                    db.execute("INSERT INTO departments(name) VALUES (?)", (department_name,))
                    dept_id = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
                else:
                    dept_id = dept["id"]
                desig = db.execute("SELECT id FROM designations WHERE lower(name)=lower(?)", (designation_name,)).fetchone()
                if not desig:
                    db.execute("INSERT INTO designations(name) VALUES (?)", (designation_name,))
                    desig_id = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
                else:
                    desig_id = desig["id"]
                manager_ref = cell(row, "manager_email") or cell(row, "manager_code")
                manager_id = None
                if manager_ref:
                    manager = db.execute("SELECT id FROM users WHERE lower(email)=lower(?) OR employee_code=?", (manager_ref, manager_ref)).fetchone()
                    manager_id = manager["id"] if manager else None
                project_ref = cell(row, "project_code") or cell(row, "project_name")
                project_id = None
                if project_ref:
                    project = db.execute("SELECT id FROM projects WHERE upper(project_code)=upper(?) OR lower(project_name)=lower(?)", (project_ref, project_ref)).fetchone()
                    project_id = project["id"] if project else None
                phone = cell(row, "phone")
                address = cell(row, "address")
                emergency_contact = cell(row, "emergency_contact")
                join_date = cell(row, "join_date", date.today().isoformat())
                monthly_basic = float(cell(row, "monthly_basic", "0") or 0)
                default_allowances = float(cell(row, "default_allowances", "0") or 0)
                deduction_per_absent = float(cell(row, "deduction_per_absent", "0") or 0)
                deduction_per_late = float(cell(row, "deduction_per_late", "0") or 0)
                is_active = 0 if cell(row, "is_active", "1").lower() in {"0", "false", "no", "inactive"} else 1
                db.execute(
                    "INSERT INTO users(full_name, email, employee_code, password_hash, role, department_id, designation_id, manager_id, project_id, phone, address, emergency_contact, join_date, monthly_basic, default_allowances, deduction_per_absent, deduction_per_late, is_active) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (full_name, email, employee_code, generate_password_hash(password), role, dept_id, desig_id, manager_id, project_id, phone, address, emergency_contact, join_date, monthly_basic, default_allowances, deduction_per_absent, deduction_per_late, is_active),
                )
                new_id = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
                for lt in db.execute("SELECT id, annual_quota FROM leave_types").fetchall():
                    db.execute("INSERT INTO leave_balances(user_id, leave_type_id, total_days, used_days, remaining_days) VALUES (?, ?, ?, 0, ?)", (new_id, lt["id"], lt["annual_quota"], lt["annual_quota"]))
                created += 1
            db.commit()
            flash(f"Bulk upload complete. Created {created} employee(s), skipped {skipped} row(s).", "success")
            return redirect(url_for("team"))
        except Exception as exc:
            flash(f"Could not process Excel file: {exc}", "danger")
            return redirect(url_for("bulk_employee_upload"))
    sample_headers = ["full_name", "email", "employee_code", "password", "role", "department", "designation", "manager_email", "project_code", "phone", "address", "emergency_contact", "join_date", "monthly_basic", "default_allowances", "deduction_per_absent", "deduction_per_late", "is_active"]
    return render_template("employee_bulk_upload.html", sample_headers=sample_headers)


@app.route("/employees/new", methods=["GET", "POST"])
@login_required
@role_required("hr", "admin", "super_admin")
def new_employee():
    return employee_form_handler()


@app.route("/employees/<int:user_id>/edit", methods=["GET", "POST"])
@login_required
@role_required("manager", "project_manager", "site_manager", "hr", "admin", "super_admin")
def edit_employee(user_id: int):
    return employee_form_handler(user_id)


def employee_form_handler(user_id: int | None = None):
    db = get_db()
    viewer = current_user()
    departments = db.execute("SELECT * FROM departments ORDER BY name").fetchall()
    designations = db.execute("SELECT * FROM designations ORDER BY name").fetchall()
    managers = db.execute("SELECT id, full_name, role, project_id FROM users WHERE role IN ('manager','project_manager','site_manager') AND is_active=1 ORDER BY full_name").fetchall()
    projects = project_choice_rows(db, viewer if is_project_scoped_role(viewer["role"]) else None)
    employee = None
    manager_assignment_only = is_project_scoped_role(viewer["role"])

    if user_id is not None:
        employee = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not employee:
            flash("Employee not found.", "danger")
            return redirect(url_for("team"))
        if manager_assignment_only:
            allowed_project_ids = {viewer["project_id"]} if viewer["project_id"] else set()
            employee_project_id = employee["project_id"]
            if employee["id"] == viewer["id"]:
                flash("You cannot change your own project assignment.", "danger")
                return redirect(url_for("employee_detail", user_id=user_id))
            if employee_project_id is not None and employee_project_id not in allowed_project_ids:
                flash("You can only manage employees from your own project or unassigned employees.", "danger")
                return redirect(url_for("team"))
            managers = [m for m in managers if m["id"] == viewer["id"] or (viewer["project_id"] and m["project_id"] == viewer["project_id"])]
    elif manager_assignment_only:
        flash("Project managers can assign or remove employees from projects through the edit employee page only.", "warning")
        return redirect(url_for("team"))

    if request.method == "POST":
        form = request.form
        manager_id = form.get("manager_id") or None
        project_id = form.get("project_id") or None

        if manager_assignment_only:
            if user_id is None or employee is None:
                flash("Invalid employee record.", "danger")
                return redirect(url_for("team"))
            normalized_project_id = int(project_id) if project_id not in (None, "", "None") else None
            if normalized_project_id not in {None, viewer["project_id"]}:
                flash("You can only assign employees to your own project or remove them from it.", "danger")
                return redirect(url_for("employee_detail", user_id=user_id))
            normalized_manager_id = int(manager_id) if manager_id not in (None, "", "None") else None
            if normalized_manager_id not in {None, viewer["id"]}:
                chosen_manager = db.execute("SELECT id, project_id FROM users WHERE id=?", (normalized_manager_id,)).fetchone()
                if not chosen_manager or chosen_manager["project_id"] != viewer["project_id"]:
                    flash("You can assign only yourself or a manager from your project.", "danger")
                    return redirect(url_for("employee_detail", user_id=user_id))
            db.execute(
                "UPDATE users SET manager_id=?, project_id=? WHERE id=?",
                (normalized_manager_id, normalized_project_id, user_id),
            )
            project_label = "Unassigned" if normalized_project_id is None else (db.execute("SELECT project_name FROM projects WHERE id=?", (normalized_project_id,)).fetchone() or {}).get("project_name", "Assigned Project")
            notify_user(user_id, "Project assignment updated", f"Your project assignment was updated to {project_label}.", url_for("employee_detail", user_id=user_id))
            log_audit("Employee", "Project Assignment Updated", f"Updated project assignment for employee {employee['employee_code']}", user_id)
            db.commit()
            flash("Employee project assignment updated successfully.", "success")
            return redirect(url_for("employee_detail", user_id=user_id))

        if user_id is None:
            db.execute(
                "INSERT INTO users(full_name, email, employee_code, password_hash, role, department_id, designation_id, manager_id, project_id, phone, address, emergency_contact, join_date, monthly_basic, default_allowances, deduction_per_absent, deduction_per_late, is_active) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (form["full_name"].strip(), form["email"].strip(), form["employee_code"].strip(), generate_password_hash(form["password"]), form["role"], form["department_id"], form["designation_id"], manager_id, project_id, form["phone"].strip(), form["address"].strip(), form["emergency_contact"].strip(), form["join_date"], float(form.get("monthly_basic") or 0), float(form.get("default_allowances") or 0), float(form.get("deduction_per_absent") or 0), float(form.get("deduction_per_late") or 0), 1 if form.get("is_active", "1") == "1" else 0),
            )
            new_id = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
            for lt in db.execute("SELECT id, annual_quota FROM leave_types").fetchall():
                db.execute("INSERT INTO leave_balances(user_id, leave_type_id, total_days, used_days, remaining_days) VALUES (?, ?, ?, 0, ?)", (new_id, lt["id"], lt["annual_quota"], lt["annual_quota"]))
            notify_user(new_id, "Account created", "Your employee portal account has been created.", url_for("profile"), "Employee portal account created", f"Welcome to the employee portal. Login email: {form['email'].strip()}")
            log_audit("Employee", "Created", f"Created employee {form['employee_code']}", new_id)
            db.commit()
            flash("Employee created successfully.", "success")
            return redirect(url_for("employee_detail", user_id=new_id))
        db.execute(
            "UPDATE users SET full_name=?, email=?, employee_code=?, role=?, department_id=?, designation_id=?, manager_id=?, project_id=?, phone=?, address=?, emergency_contact=?, join_date=?, monthly_basic=?, default_allowances=?, deduction_per_absent=?, deduction_per_late=?, is_active=? WHERE id=?",
            (form["full_name"].strip(), form["email"].strip(), form["employee_code"].strip(), form["role"], form["department_id"], form["designation_id"], manager_id, project_id, form["phone"].strip(), form["address"].strip(), form["emergency_contact"].strip(), form["join_date"], float(form.get("monthly_basic") or 0), float(form.get("default_allowances") or 0), float(form.get("deduction_per_absent") or 0), float(form.get("deduction_per_late") or 0), 1 if form.get("is_active", "1") == "1" else 0, user_id),
        )
        notify_user(user_id, "Profile updated", "Your employee profile details were updated by HR/Admin.", url_for("employee_detail", user_id=user_id))
        log_audit("Employee", "Updated", f"Updated employee {form['employee_code']}", user_id)
        db.commit()
        flash("Employee updated successfully.", "success")
        return redirect(url_for("employee_detail", user_id=user_id))
    return render_template("employee_form.html", departments=departments, designations=designations, managers=managers, employee=employee, projects=projects, role_options=get_role_options(), role_label=role_label, viewer=viewer, manager_assignment_only=manager_assignment_only)


@app.route("/employees/<int:user_id>/delete", methods=["POST"])
@login_required
@role_required("admin", "super_admin")
def delete_employee(user_id: int):
    db = get_db()
    actor = current_user()
    employee = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not employee:
        flash("Employee not found.", "danger")
        return redirect(url_for("team"))
    if actor and actor["id"] == user_id:
        flash("You cannot delete your own account while logged in.", "danger")
        return redirect(url_for("employee_detail", user_id=user_id))
    if employee["role"] in {"admin", "super_admin"}:
        active_super_admins = db.execute("SELECT COUNT(*) AS c FROM users WHERE role IN ('admin', 'super_admin') AND is_active=1").fetchone()["c"]
        if active_super_admins <= 1:
            flash("You cannot delete the last active Super Admin.", "danger")
            return redirect(url_for("employee_detail", user_id=user_id))

    avatar_filename = employee["avatar_filename"]

    db.execute("UPDATE users SET manager_id=NULL WHERE manager_id=?", (user_id,))
    db.execute("UPDATE projects SET created_by=NULL WHERE created_by=?", (user_id,))
    db.execute("UPDATE audit_logs SET actor_user_id=NULL WHERE actor_user_id=?", (user_id,))
    db.execute("UPDATE audit_logs SET target_user_id=NULL WHERE target_user_id=?", (user_id,))
    db.execute("UPDATE leave_history SET action_by=NULL WHERE action_by=?", (user_id,))
    db.execute("DELETE FROM notifications WHERE user_id=?", (user_id,))
    db.execute("DELETE FROM email_queue WHERE to_user_id=?", (user_id,))
    db.execute("DELETE FROM employee_documents WHERE user_id=?", (user_id,))
    db.execute("DELETE FROM payroll_slips WHERE user_id=?", (user_id,))
    db.execute("DELETE FROM attendance WHERE user_id=?", (user_id,))
    db.execute("DELETE FROM leave_history WHERE leave_application_id IN (SELECT id FROM leave_applications WHERE user_id=?)", (user_id,))
    db.execute("DELETE FROM leave_applications WHERE user_id=?", (user_id,))
    db.execute("DELETE FROM leave_balances WHERE user_id=?", (user_id,))
    db.execute("DELETE FROM users WHERE id=?", (user_id,))

    if avatar_filename:
        delete_stored_file(avatar_filename)

    log_audit("Employee", "Deleted", f"Deleted employee {employee['employee_code']}", None)
    db.commit()
    flash("Employee deleted successfully.", "success")
    return redirect(url_for("team"))


@app.route("/employees/<int:user_id>/reset-password", methods=["POST"])
@login_required
@role_required("admin", "super_admin", "hr")
def reset_employee_password(user_id: int):
    db = get_db()
    employee = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not employee:
        flash("Employee not found.", "danger")
        return redirect(url_for("team"))
    temp_password = f"Reset@{secrets.token_hex(4)}A1"
    db.execute("UPDATE users SET password_hash=? WHERE id=?", (generate_password_hash(temp_password), user_id))
    notify_user(
        user_id,
        "Password reset",
        "HR/Admin generated a temporary password for your account.",
        url_for("change_password"),
        "Password reset",
        f"Your temporary password is: {temp_password}\nPlease sign in and change it immediately.",
    )
    log_audit("Security", "Password Reset", "Temporary password generated by HR/Admin", user_id)
    db.commit()
    flash("Temporary password generated and queued in the email center.", "success")
    return redirect(url_for("employee_detail", user_id=user_id))


@app.route("/attendance")
@login_required
def attendance_view():
    user = current_user()
    db = get_db()
    selected_user_id = request.args.get("user_id", type=int) or user["id"]
    month = request.args.get("month") or date.today().strftime("%Y-%m")
    if user["role"] == "employee":
        selected_user_id = user["id"]
    else:
        allowed_condition, allowed_params = visible_user_filter(user, "u")
        allowed = db.execute(f"SELECT COUNT(*) AS c FROM users u WHERE u.id=? AND {allowed_condition}", (selected_user_id, *allowed_params)).fetchone()["c"]
        if selected_user_id != user["id"] and not allowed:
            selected_user_id = user["id"]
    rows = db.execute(
        "SELECT a.*, hc.title AS holiday_title, hc.holiday_type FROM attendance a LEFT JOIN holiday_calendar hc ON a.attendance_date=hc.holiday_date WHERE a.user_id=? AND substr(a.attendance_date,1,7)=? ORDER BY a.attendance_date DESC",
        (selected_user_id, month),
    ).fetchall()
    employees = []
    if can_manage_people(user["role"]):
        employees = team_user_rows(db, user)
    summary = {
        "present": sum(1 for r in rows if r["status"] == "Present"),
        "absent": sum(1 for r in rows if r["status"] == "Absent"),
        "late": sum(1 for r in rows if r["status"] == "Late"),
        "hours": round(sum(r["hours_worked"] for r in rows), 1),
        "ot_hours": round(sum((r["ot_hours"] or 0) for r in rows), 1),
    }
    selected_employee = db.execute("SELECT id, full_name FROM users WHERE id=?", (selected_user_id,)).fetchone()
    holiday_rows = db.execute("SELECT * FROM holiday_calendar WHERE substr(holiday_date,1,7)=? ORDER BY holiday_date", (month,)).fetchall()
    return render_template("attendance.html", rows=rows, employees=employees, selected_user_id=selected_user_id, month=month, summary=summary, selected_employee=selected_employee, holiday_rows=holiday_rows)


@app.route("/payroll")
@login_required
def payroll_view():
    user = current_user()
    db = get_db()
    selected_user_id = request.args.get("user_id", type=int) or user["id"]
    if user["role"] == "employee":
        selected_user_id = user["id"]
    else:
        allowed_condition, allowed_params = visible_user_filter(user, "u")
        allowed = db.execute(f"SELECT COUNT(*) AS c FROM users u WHERE u.id=? AND {allowed_condition}", (selected_user_id, *allowed_params)).fetchone()["c"]
        if selected_user_id != user["id"] and not allowed:
            selected_user_id = user["id"]
    slips = db.execute("SELECT * FROM payroll_slips WHERE user_id=? ORDER BY id DESC", (selected_user_id,)).fetchall()
    employees = team_user_rows(db, user) if can_manage_people(user["role"]) else []
    selected_employee = db.execute("SELECT id, full_name FROM users WHERE id=?", (selected_user_id,)).fetchone()
    return render_template("payroll.html", slips=slips, employees=employees, selected_user_id=selected_user_id, selected_employee=selected_employee)




@app.route("/attendance/add", methods=["GET", "POST"])
@login_required
@role_required("admin", "super_admin")
def attendance_add():
    db = get_db()
    employees = db.execute("SELECT id, full_name FROM users WHERE is_active=1 ORDER BY full_name").fetchall()
    statuses = ["Present", "Absent", "Leave", "Late", "Half Day", "Holiday", "Vacation"]
    if request.method == "POST":
        user_id = request.form.get("user_id", type=int)
        attendance_date = (request.form.get("attendance_date") or "").strip()
        check_in = (request.form.get("check_in") or "").strip() or None
        check_out = (request.form.get("check_out") or "").strip() or None
        status = (request.form.get("status") or "Present").strip()
        remarks = (request.form.get("remarks") or "").strip() or None
        hours_worked = request.form.get("hours_worked", type=float)
        if hours_worked is None:
            hours_worked = calculate_hours(check_in, check_out)
        manual_ot = request.form.get("ot_hours", type=float)
        ot_hours = compute_ot_hours(db, attendance_date, status, hours_worked, manual_ot)
        if not user_id or not attendance_date:
            flash("Employee and date are required.", "danger")
        else:
            db.execute(
                "INSERT INTO attendance(user_id, attendance_date, check_in, check_out, status, hours_worked, ot_hours, remarks) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (user_id, attendance_date, check_in, check_out, status, hours_worked, ot_hours, remarks),
            )
            employee = db.execute("SELECT full_name FROM users WHERE id=?", (user_id,)).fetchone()
            log_audit("Attendance", "Added", f"Added attendance for {employee['full_name']} on {attendance_date}", user_id)
            db.commit()
            flash("Attendance record added.", "success")
            return redirect(url_for("attendance_view", user_id=user_id, month=attendance_date[:7]))
    return render_template("attendance_form.html", record=None, employees=employees, statuses=statuses)


@app.route("/attendance/edit/<int:attendance_id>", methods=["GET", "POST"])
@login_required
@role_required("admin", "super_admin")
def attendance_edit(attendance_id: int):
    db = get_db()
    record = db.execute("SELECT * FROM attendance WHERE id=?", (attendance_id,)).fetchone()
    if not record:
        flash("Attendance record not found.", "danger")
        return redirect(url_for("attendance_view"))
    employees = db.execute("SELECT id, full_name FROM users WHERE is_active=1 ORDER BY full_name").fetchall()
    statuses = ["Present", "Absent", "Leave", "Late", "Half Day", "Holiday", "Vacation"]
    if request.method == "POST":
        user_id = request.form.get("user_id", type=int)
        attendance_date = (request.form.get("attendance_date") or "").strip()
        check_in = (request.form.get("check_in") or "").strip() or None
        check_out = (request.form.get("check_out") or "").strip() or None
        status = (request.form.get("status") or "Present").strip()
        remarks = (request.form.get("remarks") or "").strip() or None
        hours_worked = request.form.get("hours_worked", type=float)
        if hours_worked is None:
            hours_worked = calculate_hours(check_in, check_out)
        manual_ot = request.form.get("ot_hours", type=float)
        ot_hours = compute_ot_hours(db, attendance_date, status, hours_worked, manual_ot)
        db.execute(
            "UPDATE attendance SET user_id=?, attendance_date=?, check_in=?, check_out=?, status=?, hours_worked=?, ot_hours=?, remarks=? WHERE id=?",
            (user_id, attendance_date, check_in, check_out, status, hours_worked, ot_hours, remarks, attendance_id),
        )
        employee = db.execute("SELECT full_name FROM users WHERE id=?", (user_id,)).fetchone()
        log_audit("Attendance", "Edited", f"Edited attendance for {employee['full_name']} on {attendance_date}", user_id)
        db.commit()
        flash("Attendance record updated.", "success")
        return redirect(url_for("attendance_view", user_id=user_id, month=attendance_date[:7]))
    return render_template("attendance_form.html", record=record, employees=employees, statuses=statuses)


@app.route("/attendance/delete/<int:attendance_id>", methods=["POST"])
@login_required
@role_required("admin", "super_admin")
def attendance_delete(attendance_id: int):
    db = get_db()
    record = db.execute("SELECT * FROM attendance WHERE id=?", (attendance_id,)).fetchone()
    if not record:
        flash("Attendance record not found.", "danger")
        return redirect(url_for("attendance_view"))
    db.execute("DELETE FROM attendance WHERE id=?", (attendance_id,))
    log_audit("Attendance", "Deleted", f"Deleted attendance entry on {record['attendance_date']}", record['user_id'])
    db.commit()
    flash("Attendance record deleted.", "success")
    return redirect(url_for("attendance_view", user_id=record['user_id'], month=str(record['attendance_date'])[:7]))


@app.route("/payroll/add", methods=["GET", "POST"])
@login_required
@role_required("admin", "super_admin")
def payroll_add():
    db = get_db()
    employees = db.execute("SELECT id, full_name FROM users WHERE is_active=1 ORDER BY full_name").fetchall()
    if request.method == "POST":
        user_id = request.form.get("user_id", type=int)
        month_label = (request.form.get("month_label") or "").strip()
        basic_salary = request.form.get("basic_salary", type=float) or 0.0
        allowances = request.form.get("allowances", type=float) or 0.0
        deductions = request.form.get("deductions", type=float) or 0.0
        net_salary = basic_salary + allowances - deductions
        if not user_id or not month_label:
            flash("Employee and month are required.", "danger")
        else:
            db.execute(
                "INSERT INTO payroll_slips(user_id, month_label, basic_salary, allowances, deductions, net_salary, generated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (user_id, month_label, basic_salary, allowances, deductions, net_salary, now_str()),
            )
            employee = db.execute("SELECT full_name FROM users WHERE id=?", (user_id,)).fetchone()
            log_audit("Payroll", "Added", f"Added payslip for {employee['full_name']} ({month_label})", user_id)
            db.commit()
            flash("Payslip added.", "success")
            return redirect(url_for("payroll_view", user_id=user_id))
    return render_template("payroll_form.html", slip=None, employees=employees)


@app.route("/payroll/edit/<int:slip_id>", methods=["GET", "POST"])
@login_required
@role_required("admin", "super_admin")
def payroll_edit(slip_id: int):
    db = get_db()
    slip = db.execute("SELECT * FROM payroll_slips WHERE id=?", (slip_id,)).fetchone()
    if not slip:
        flash("Payslip not found.", "danger")
        return redirect(url_for("payroll_view"))
    employees = db.execute("SELECT id, full_name FROM users WHERE is_active=1 ORDER BY full_name").fetchall()
    if request.method == "POST":
        user_id = request.form.get("user_id", type=int)
        month_label = (request.form.get("month_label") or "").strip()
        basic_salary = request.form.get("basic_salary", type=float) or 0.0
        allowances = request.form.get("allowances", type=float) or 0.0
        deductions = request.form.get("deductions", type=float) or 0.0
        net_salary = basic_salary + allowances - deductions
        db.execute(
            "UPDATE payroll_slips SET user_id=?, month_label=?, basic_salary=?, allowances=?, deductions=?, net_salary=? WHERE id=?",
            (user_id, month_label, basic_salary, allowances, deductions, net_salary, slip_id),
        )
        employee = db.execute("SELECT full_name FROM users WHERE id=?", (user_id,)).fetchone()
        log_audit("Payroll", "Edited", f"Edited payslip for {employee['full_name']} ({month_label})", user_id)
        db.commit()
        flash("Payslip updated.", "success")
        return redirect(url_for("payroll_view", user_id=user_id))
    return render_template("payroll_form.html", slip=slip, employees=employees)


@app.route("/payroll/delete/<int:slip_id>", methods=["POST"])
@login_required
@role_required("admin", "super_admin")
def payroll_delete(slip_id: int):
    db = get_db()
    slip = db.execute("SELECT * FROM payroll_slips WHERE id=?", (slip_id,)).fetchone()
    if not slip:
        flash("Payslip not found.", "danger")
        return redirect(url_for("payroll_view"))
    db.execute("DELETE FROM payroll_slips WHERE id=?", (slip_id,))
    log_audit("Payroll", "Deleted", f"Deleted payslip {slip['month_label']}", slip['user_id'])
    db.commit()
    flash("Payslip deleted.", "success")
    return redirect(url_for("payroll_view", user_id=slip['user_id']))

@app.route("/reports")
@login_required
@role_required("hr", "admin", "super_admin")
def reports():
    db = get_db()
    leave_summary = db.execute("SELECT status, COUNT(*) AS total FROM leave_applications GROUP BY status ORDER BY total DESC").fetchall()
    dept_summary = db.execute("SELECT d.name AS department_name, COUNT(u.id) AS total FROM departments d LEFT JOIN users u ON u.department_id=d.id AND u.is_active=1 GROUP BY d.id, d.name ORDER BY d.name").fetchall()
    attendance_summary = db.execute("SELECT status, COUNT(*) AS total FROM attendance GROUP BY status ORDER BY total DESC").fetchall()
    return render_template("reports.html", leave_summary=leave_summary, dept_summary=dept_summary, attendance_summary=attendance_summary)


@app.route("/documents")
@login_required
def documents():
    user = current_user()
    db = get_db()
    employees = []
    if is_admin_role(user["role"]) or is_hr_role(user["role"]):
        docs = db.execute("SELECT ed.*, u.full_name FROM employee_documents ed JOIN users u ON ed.user_id=u.id ORDER BY ed.id DESC").fetchall()
        employees = db.execute("SELECT id, full_name, employee_code FROM users WHERE is_active=1 ORDER BY full_name").fetchall()
    elif is_project_scoped_role(user["role"]):
        docs = db.execute("SELECT ed.*, u.full_name FROM employee_documents ed JOIN users u ON ed.user_id=u.id WHERE u.project_id=? ORDER BY ed.id DESC", (user["project_id"],)).fetchall() if user["project_id"] else []
        employees = db.execute("SELECT id, full_name, employee_code FROM users WHERE is_active=1 AND project_id=? ORDER BY full_name", (user["project_id"],)).fetchall() if user["project_id"] else []
    else:
        docs = db.execute("SELECT ed.*, u.full_name FROM employee_documents ed JOIN users u ON ed.user_id=u.id WHERE user_id=? ORDER BY ed.id DESC", (user["id"],)).fetchall()
    return render_template("documents.html", docs=docs, employees=employees)


@app.route("/documents/upload", methods=["POST"])
@login_required
@role_required("hr", "admin", "super_admin")
def upload_document():
    db = get_db()
    user_id = int(request.form["user_id"])
    title = request.form["title"].strip()
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Please choose a file.", "danger")
        return redirect(url_for("documents"))
    if not allowed_file(file.filename):
        flash("Unsupported file type.", "danger")
        return redirect(url_for("documents"))
    filename = f"doc_{datetime.now().strftime('%Y%m%d%H%M%S')}_{secure_filename(file.filename)}"
    stored_file = upload_file_storage(file, "employee_portal/documents", ALLOWED_EXTENSIONS)
    db.execute("INSERT INTO employee_documents(user_id, title, file_name, uploaded_at) VALUES (?, ?, ?, ?)", (user_id, title, stored_file, now_str()))
    notify_user(user_id, "New document uploaded", f"A new document titled '{title}' has been uploaded to your portal.", url_for("documents"))
    log_audit("Documents", "Uploaded", f"Uploaded document {title}", user_id)
    db.commit()
    flash("Document uploaded successfully.", "success")
    return redirect(url_for("documents"))


@app.route("/notifications")
@login_required
def notifications_view():
    db = get_db()
    user = current_user()
    mark_all = request.args.get("mark_all")
    if mark_all == "1":
        db.execute("UPDATE notifications SET is_read=1 WHERE user_id=?", (user["id"],))
        db.commit()
        flash("All notifications marked as read.", "success")
        return redirect(url_for("notifications_view"))
    notifications = db.execute("SELECT * FROM notifications WHERE user_id=? ORDER BY id DESC", (user["id"],)).fetchall()
    return render_template("notifications.html", notifications=notifications)


@app.route("/notifications/<int:notification_id>/read")
@login_required
def mark_notification_read(notification_id: int):
    db = get_db()
    user = current_user()
    row = db.execute("SELECT * FROM notifications WHERE id=? AND user_id=?", (notification_id, user["id"])).fetchone()
    if row:
        db.execute("UPDATE notifications SET is_read=1 WHERE id=?", (notification_id,))
        db.commit()
        if row["link"]:
            return redirect(row["link"])
    return redirect(url_for("notifications_view"))


@app.route("/email-center")
@login_required
@role_required("admin")
def email_center():
    emails = get_db().execute(
        "SELECT eq.*, u.full_name FROM email_queue eq LEFT JOIN users u ON eq.to_user_id=u.id ORDER BY eq.id DESC LIMIT 200"
    ).fetchall()
    return render_template("email_center.html", emails=emails)


@app.route("/email-center/<int:email_id>/sent", methods=["POST"])
@login_required
@role_required("admin")
def mark_email_sent(email_id: int):
    db = get_db()
    db.execute("UPDATE email_queue SET status='Sent' WHERE id=?", (email_id,))
    log_audit("Email", "Marked Sent", f"Marked email #{email_id} as sent")
    db.commit()
    flash("Email marked as sent.", "success")
    return redirect(url_for("email_center"))


@app.route("/payroll/bulk-upload", methods=["GET", "POST"])
@login_required
@role_required("admin")
def payroll_bulk_upload():
    db = get_db()
    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename:
            flash("Please choose an Excel file.", "danger")
            return redirect(url_for("payroll_bulk_upload"))
        try:
            wb = load_workbook(io.BytesIO(file.read()), data_only=True)
            ws = wb.active
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                employee_code = str(row[0]).strip()
                month_label = str(row[1]).strip()
                basic_salary = float(row[2] or 0)
                allowances = float(row[3] or 0)
                deductions = float(row[4] or 0)
                user = db.execute("SELECT id FROM users WHERE employee_code=?", (employee_code,)).fetchone()
                if not user:
                    continue
                net_salary = round(basic_salary + allowances - deductions, 2)
                existing = db.execute("SELECT id FROM payroll_slips WHERE user_id=? AND month_label=?", (user["id"], month_label)).fetchone()
                if existing:
                    db.execute("UPDATE payroll_slips SET basic_salary=?, allowances=?, deductions=?, net_salary=?, generated_at=? WHERE id=?", (basic_salary, allowances, deductions, net_salary, now_str(), existing["id"]))
                else:
                    db.execute("INSERT INTO payroll_slips(user_id, month_label, basic_salary, allowances, deductions, net_salary, generated_at) VALUES (?, ?, ?, ?, ?, ?, ?)", (user["id"], month_label, basic_salary, allowances, deductions, net_salary, now_str()))
                count += 1
            db.commit()
            flash(f"Payslips imported/updated for {count} rows.", "success")
            return redirect(url_for("payroll_view"))
        except Exception as exc:
            flash(f"Upload failed: {exc}", "danger")
    return render_template("payroll_bulk_upload.html")


@app.route("/payroll/generate-auto", methods=["GET", "POST"])
@login_required
@role_required("admin")
def payroll_generate_auto():
    db = get_db()
    employees = db.execute("SELECT id, full_name, employee_code FROM users WHERE is_active=1 ORDER BY full_name").fetchall()
    month = request.form.get("month") or date.today().strftime("%Y-%m")
    if request.method == "POST":
        user_id = request.form.get("user_id", type=int)
        if user_id:
            upsert_payroll_from_attendance(user_id, month)
        else:
            for emp in employees:
                upsert_payroll_from_attendance(emp["id"], month)
        db.commit()
        flash("Payslip generation completed.", "success")
        return redirect(url_for("payroll_view"))
    return render_template("payroll_generate_auto.html", employees=employees, month=month)


@app.route("/attendance/bulk-upload", methods=["GET", "POST"])
@login_required
@role_required("admin")
def attendance_bulk_upload():
    db = get_db()
    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename:
            flash("Please choose an Excel file.", "danger")
            return redirect(url_for("attendance_bulk_upload"))
        try:
            wb = load_workbook(io.BytesIO(file.read()), data_only=True)
            ws = wb.active
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0] or not row[1]:
                    continue
                employee_code = str(row[0]).strip()
                attendance_date = str(row[1]).strip()
                check_in = str(row[2]).strip() if len(row) > 2 and row[2] not in (None, "") else None
                check_out = str(row[3]).strip() if len(row) > 3 and row[3] not in (None, "") else None
                status = str(row[4]).strip() if len(row) > 4 and row[4] not in (None, "") else "Present"
                remarks = str(row[5]).strip() if len(row) > 5 and row[5] not in (None, "") else None
                manual_ot = float(row[6]) if len(row) > 6 and row[6] not in (None, "") else None
                user = db.execute("SELECT id FROM users WHERE employee_code=?", (employee_code,)).fetchone()
                if not user:
                    continue
                hours_worked = calculate_hours(check_in, check_out)
                ot_hours = compute_ot_hours(db, attendance_date, status, hours_worked, manual_ot)
                existing = db.execute("SELECT id FROM attendance WHERE user_id=? AND attendance_date=?", (user["id"], attendance_date)).fetchone()
                if existing:
                    db.execute("UPDATE attendance SET check_in=?, check_out=?, status=?, hours_worked=?, ot_hours=?, remarks=? WHERE id=?", (check_in, check_out, status, hours_worked, ot_hours, remarks, existing["id"]))
                else:
                    db.execute("INSERT INTO attendance(user_id, attendance_date, check_in, check_out, status, hours_worked, ot_hours, remarks) VALUES (?, ?, ?, ?, ?, ?, ?, ?)", (user["id"], attendance_date, check_in, check_out, status, hours_worked, ot_hours, remarks))
                count += 1
            db.commit()
            flash(f"Attendance imported/updated for {count} rows.", "success")
            return redirect(url_for("attendance_view"))
        except Exception as exc:
            flash(f"Upload failed: {exc}", "danger")
    return render_template("attendance_bulk_upload.html")


@app.route("/attendance/monthly-editor", methods=["GET", "POST"])
@login_required
@role_required("admin")
def attendance_monthly_editor():
    db = get_db()
    attendance_date = request.values.get("attendance_date") or date.today().strftime("%Y-%m-%d")
    employees = db.execute("SELECT id, full_name, employee_code FROM users WHERE is_active=1 ORDER BY full_name").fetchall()
    statuses = ["Present", "Absent", "Leave", "Late", "Half Day", "Holiday", "Vacation"]
    existing_rows = {r["user_id"]: r for r in db.execute("SELECT * FROM attendance WHERE attendance_date=?", (attendance_date,)).fetchall()}
    if request.method == "POST":
        for emp in employees:
            prefix = f"emp_{emp['id']}_"
            status = request.form.get(prefix + "status", "Present")
            check_in = (request.form.get(prefix + "check_in") or "").strip() or None
            check_out = (request.form.get(prefix + "check_out") or "").strip() or None
            remarks = (request.form.get(prefix + "remarks") or "").strip() or None
            hours = calculate_hours(check_in, check_out)
            ot_hours = compute_ot_hours(db, attendance_date, status, hours)
            current = existing_rows.get(emp["id"])
            if current:
                db.execute("UPDATE attendance SET status=?, check_in=?, check_out=?, hours_worked=?, ot_hours=?, remarks=? WHERE id=?", (status, check_in, check_out, hours, ot_hours, remarks, current["id"]))
            else:
                db.execute("INSERT INTO attendance(user_id, attendance_date, check_in, check_out, status, hours_worked, ot_hours, remarks) VALUES (?, ?, ?, ?, ?, ?, ?, ?)", (emp["id"], attendance_date, check_in, check_out, status, hours, ot_hours, remarks))
        db.commit()
        flash("Attendance sheet saved successfully.", "success")
        return redirect(url_for("attendance_monthly_editor", attendance_date=attendance_date))
    holiday_row = get_holiday_row(db, attendance_date)
    return render_template("attendance_monthly_editor.html", employees=employees, existing=existing_rows, statuses=statuses, attendance_date=attendance_date, holiday_row=holiday_row)


@app.route("/masters", methods=["GET", "POST"])
@login_required
@role_required("admin")
def masters_view():
    db = get_db()
    if request.method == "POST":
        action = request.form.get("action")
        try:
            if action == "department":
                name = (request.form.get("name") or "").strip()
                if name:
                    db.execute("INSERT INTO departments(name) VALUES (?)", (name,))
                    flash("Department added.", "success")
            elif action == "designation":
                name = (request.form.get("name") or "").strip()
                if name:
                    db.execute("INSERT INTO designations(name) VALUES (?)", (name,))
                    flash("Designation added.", "success")
            elif action == "promote_manager":
                user_id = request.form.get("user_id", type=int)
                if user_id:
                    db.execute("UPDATE users SET role='manager' WHERE id=?", (user_id,))
                    flash("User promoted to manager.", "success")
            db.commit()
        except sqlite3.IntegrityError:
            flash("This value already exists.", "warning")
        return redirect(url_for("masters_view"))
    departments = db.execute("SELECT * FROM departments ORDER BY name").fetchall()
    designations = db.execute("SELECT * FROM designations ORDER BY name").fetchall()
    managers = db.execute("SELECT id, full_name, employee_code FROM users WHERE role='manager' AND is_active=1 ORDER BY full_name").fetchall()
    non_managers = db.execute("SELECT id, full_name, employee_code FROM users WHERE role!='manager' AND is_active=1 ORDER BY full_name").fetchall()
    return render_template("masters.html", departments=departments, designations=designations, managers=managers, non_managers=non_managers, role_options=get_role_options())


@app.route("/calendar", methods=["GET", "POST"])
@login_required
@role_required("admin")
def calendar_view():
    db = get_db()
    if request.method == "POST":
        holiday_date = (request.form.get("holiday_date") or "").strip()
        title = (request.form.get("title") or "").strip()
        holiday_type = (request.form.get("holiday_type") or "Holiday").strip()
        if holiday_date and title:
            existing = db.execute("SELECT id FROM holiday_calendar WHERE holiday_date=?", (holiday_date,)).fetchone()
            if existing:
                db.execute("UPDATE holiday_calendar SET title=?, holiday_type=? WHERE id=?", (title, holiday_type, existing["id"]))
                flash("Calendar day updated.", "success")
            else:
                db.execute("INSERT INTO holiday_calendar(holiday_date, title, holiday_type, created_at) VALUES (?, ?, ?, ?)", (holiday_date, title, holiday_type, now_str()))
                flash("Calendar day added.", "success")
            db.commit()
        else:
            flash("Date and title are required.", "danger")
        return redirect(url_for("calendar_view"))
    month = request.args.get("month") or date.today().strftime("%Y-%m")
    rows = db.execute("SELECT * FROM holiday_calendar WHERE substr(holiday_date,1,7)=? ORDER BY holiday_date", (month,)).fetchall()
    return render_template("calendar.html", rows=rows, month=month)


@app.route("/calendar/delete/<int:holiday_id>", methods=["POST"])
@login_required
@role_required("admin")
def calendar_delete(holiday_id: int):
    db = get_db()
    db.execute("DELETE FROM holiday_calendar WHERE id=?", (holiday_id,))
    db.commit()
    flash("Calendar day removed.", "success")
    return redirect(url_for("calendar_view"))


@app.route("/settings", methods=["GET", "POST"])
@login_required
@role_required("admin")
def settings_view():
    db = get_db()
    current = db.execute("SELECT * FROM company_settings ORDER BY id DESC LIMIT 1").fetchone()
    if request.method == "POST":
        db.execute(
            "UPDATE company_settings SET company_name=?, leave_workflow=?, default_working_hours=?, allow_document_upload=? WHERE id=?",
            (request.form["company_name"].strip(), request.form["leave_workflow"].strip(), float(request.form["default_working_hours"]), 1 if request.form.get("allow_document_upload") else 0, current["id"]),
        )
        log_audit("Settings", "Updated", "Updated company settings")
        db.commit()
        flash("Settings updated successfully.", "success")
        return redirect(url_for("settings_view"))
    recent_logs = db.execute(
        "SELECT al.*, a.full_name AS actor_name, t.full_name AS target_name FROM audit_logs al LEFT JOIN users a ON al.actor_user_id=a.id LEFT JOIN users t ON al.target_user_id=t.id ORDER BY al.id DESC LIMIT 20"
    ).fetchall()
    return render_template("settings.html", setting=current, recent_logs=recent_logs)


@app.route("/uploads/<path:filename>")
@login_required
def uploaded_file(filename: str):
    if is_external_file(filename):
        return redirect(filename)
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


@app.route("/initdb")
def initialize_database():
    initialize_postgres()
    return "Database initialized successfully."


initialize_postgres()


if __name__ == "__main__":
    app.run(debug=True)
